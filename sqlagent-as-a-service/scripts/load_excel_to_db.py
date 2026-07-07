"""Load the FAB POC Excel workbook into a relational database (MySQL or PostgreSQL).

Tables are created from the semantic layer (sql_agent/semantic_layer/schema.yaml), so
column names, types and primary keys stay in lock-step with everything else the agent
uses. The same script works on MySQL and Postgres — only the DSN changes.

IMPORTANT: this is a WRITE/DDL operation. Use an admin (or write-capable) credential
here, NOT the SELECT-only credential the agent runs with.

Usage
-----
    # MySQL
    uv run python scripts/load_excel_to_db.py \
        --dsn "mysql+pymysql://admin:pw@localhost:3306/fab_pricing" --drop

    # PostgreSQL
    uv run python scripts/load_excel_to_db.py \
        --dsn "postgresql+psycopg2://admin:pw@localhost:5432/fab_pricing" --drop

    # Dry run against throwaway SQLite (no server needed) to validate mapping
    uv run python scripts/load_excel_to_db.py --dsn "sqlite://" --drop

Flags
-----
    --excel PATH   workbook to load (default: the POC file under documentation/)
    --dsn   DSN    SQLAlchemy URL of the target DB (or set LOAD_DSN env var)
    --drop         drop & recreate the five tables before loading
"""

from __future__ import annotations

import argparse
import datetime as dt
import os
import sys
from pathlib import Path

import openpyxl
from sqlalchemy import (
    BigInteger,
    Date,
    Float,
    MetaData,
    String,
    Table,
    create_engine,
    insert,
)
from sqlalchemy import Column as SAColumn

# Make `sql_agent` importable when run as a standalone script.
sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from sql_agent.semantic_layer.loader import load_semantic_layer  # noqa: E402

DEFAULT_EXCEL = (
    Path(__file__).resolve().parents[1]
    / "documentation"
    / "fab_pricing_recommendation_poc_synthetic_data.xlsx"
)

# Map a semantic-layer column type to a portable SQLAlchemy column type.
_TYPE_MAP = {
    "str": lambda: String(255),
    "enum": lambda: String(64),
    "int": BigInteger,
    "float": Float,
    "date": Date,
}


def _sa_type(col_type: str):
    factory = _TYPE_MAP.get(col_type, lambda: String(255))
    return factory() if callable(factory) else factory


def build_tables(metadata: MetaData) -> dict[str, Table]:
    """Build SQLAlchemy Table objects from the semantic layer (the source of truth)."""
    layer = load_semantic_layer()
    tables: dict[str, Table] = {}
    for tname, tdef in layer.tables.items():
        columns = [
            SAColumn(
                col.name,
                _sa_type(col.type),
                primary_key=(col.name == tdef.primary_key),
            )
            for col in tdef.columns.values()
        ]
        tables[tname] = Table(tname, metadata, *columns)
    return tables


def _coerce(value, sa_type):
    """Normalise an Excel cell value for the target column type."""
    if value is None or (isinstance(value, str) and value.strip() == ""):
        return None
    if isinstance(sa_type, Date):
        if isinstance(value, dt.datetime):
            return value.date()
        if isinstance(value, dt.date):
            return value
        # Excel often stores dates as plain strings, e.g. "2026-06-01".
        return dt.date.fromisoformat(str(value).strip()[:10])
    if isinstance(sa_type, (Float, BigInteger)) and isinstance(value, str):
        num = float(value)
        return int(num) if isinstance(sa_type, BigInteger) else num
    if isinstance(sa_type, String) and not isinstance(value, str):
        return str(value)
    return value


def load(excel_path: Path, dsn: str, drop: bool) -> None:
    metadata = MetaData()
    tables = build_tables(metadata)

    engine = create_engine(dsn, future=True)
    if drop:
        metadata.drop_all(engine)
    metadata.create_all(engine)
    print(f"Tables ready in {engine.url.get_backend_name()}: {', '.join(tables)}")

    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)

    total = 0
    with engine.begin() as conn:
        for tname, table in tables.items():
            if tname not in wb.sheetnames:
                print(f"  ! sheet '{tname}' missing in workbook — skipped")
                continue
            ws = wb[tname]
            rows_iter = ws.iter_rows(values_only=True)
            header = [h for h in next(rows_iter, []) if h is not None]
            col_types = {c.name: c.type for c in table.columns}

            payload = []
            for raw in rows_iter:
                if raw is None or all(v is None for v in raw):
                    continue
                record = {}
                for key, val in zip(header, raw):
                    if key in col_types:
                        record[key] = _coerce(val, col_types[key])
                if record:
                    payload.append(record)

            if payload:
                conn.execute(insert(table), payload)
            total += len(payload)
            print(f"  + {tname}: {len(payload)} rows")

    print(f"Done. {total} rows loaded into {excel_path.name}'s target DB.")


def main() -> None:
    parser = argparse.ArgumentParser(description="Load the FAB POC workbook into a DB.")
    parser.add_argument("--excel", type=Path, default=DEFAULT_EXCEL)
    parser.add_argument("--dsn", default=os.environ.get("LOAD_DSN", ""))
    parser.add_argument("--drop", action="store_true",
                        help="drop & recreate the tables before loading")
    args = parser.parse_args()

    if not args.dsn:
        parser.error("provide --dsn (or set LOAD_DSN). Use a WRITE-capable credential.")
    if not args.excel.exists():
        parser.error(f"workbook not found: {args.excel}")

    load(args.excel, args.dsn, args.drop)


if __name__ == "__main__":
    main()
